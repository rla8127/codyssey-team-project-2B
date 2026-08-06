"""데이터 내보내기.

요구사항 8번: CSV / JSONL / Excel 중 최소 2개를 지원한다. 여기서는 3개 모두 구현한다.
pandas 없이 표준 라이브러리 csv + openpyxl만 사용한다.
"""

import csv
import json
from datetime import datetime
from pathlib import Path

from src.common.logger import get_logger

logger = get_logger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent.parent
EXPORT_DIR = BASE_DIR / "output" / "exports"

# 내보낼 컬럼과 순서 (본문 content는 너무 길어 제외하고 summary만 넣는다).
COLUMNS = [
    "id",
    "category",
    "title",
    "summary",
    "sentiment",
    "status",
    "source_name",
    "collect_method",
    "published_at",
    "collected_at",
    "source_url",
]


def export_rows(rows: list[dict], fmt: str) -> Path | None:
    """rows를 지정 포맷으로 저장하고 경로를 반환한다. 대상이 없으면 None."""
    if not rows:
        logger.warning("내보낼 데이터가 없습니다.")
        return None

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = EXPORT_DIR / f"news_{stamp}.{'xlsx' if fmt == 'excel' else fmt}"

    if fmt == "csv":
        _to_csv(rows, path)
    elif fmt == "jsonl":
        _to_jsonl(rows, path)
    elif fmt == "excel":
        _to_excel(rows, path)
    else:
        raise ValueError(f"지원하지 않는 포맷: {fmt}")

    logger.info("내보내기 완료: %s (%d건)", path, len(rows))
    return path


def _pick(row: dict) -> dict:
    """COLUMNS에 있는 필드만 골라낸다 (없는 키는 빈 문자열)."""
    return {col: row.get(col, "") for col in COLUMNS}


def _to_csv(rows: list[dict], path: Path) -> None:
    # newline=""은 csv 모듈 권장값 (Windows에서 빈 줄이 끼는 것을 막는다).
    # utf-8-sig(BOM)로 저장해야 Excel에서 열 때 한글이 깨지지 않는다.
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(_pick(row))


def _to_jsonl(rows: list[dict], path: Path) -> None:
    # JSONL은 한 줄에 JSON 하나. 대용량도 줄 단위로 읽을 수 있어 로그성 데이터에 적합하다.
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(_pick(row), ensure_ascii=False) + "\n")


def _to_excel(rows: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "news"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        picked = _pick(row)
        # None은 셀에 그대로 쓰면 비어 보이지만, 문자열로 통일해 두는 편이 안전하다.
        ws.append([("" if picked[col] is None else picked[col]) for col in COLUMNS])

    # 컬럼 폭을 내용 길이에 맞춰 대략 조정한다 (너무 넓어지지 않게 상한을 둔다).
    for idx, col in enumerate(COLUMNS, start=1):
        longest = max([len(str(r.get(col) or "")) for r in rows] + [len(col)])
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(longest + 2, 50)

    # 첫 행 고정 + 자동 필터로 엑셀에서 바로 정렬/필터할 수 있게 한다.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")

    wb.save(path)
